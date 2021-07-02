# Install openpyxl, pandas, XlsxWriter in your environment or using terminal
# importing the modules
import numpy as np
import pandas as pd
from openpyxl.styles import Border, Side
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
# Import `load_workbook` module from `openpyxl`
from openpyxl import load_workbook

###############################################################################
# The next section in this script needs to be assigned before running the code
###############################################################################

# dimensions of the plate
# given below for a 384 well plate (16x24)
total_rows = 16
total_columns = 24

# number of reagents in the experiment
reagents = 3

######################################################################
# Actual Code starts: user need not change any part after this
######################################################################

def create_file(rows, columns, total_reagents):
    # major and minor axes for the row and column indexes of the plate respectively
    minor_axis = []
    for i in range(1, columns + 1):
        minor_axis.append(i)
    major_axis = []
    alpha = 'A'
    beta = ''
    for i in range(1, rows + 1):
        if i==27:
            alpha = 'A'
            beta = 'A'
        elif i>27:
            if i%26 ==1:
                alpha = 'A'
                beta = chr(ord(beta) + 1)
        major_axis.append(beta + alpha)
        alpha = chr(ord(alpha) + 1)

    # pandas DataFrame of the well numbers(well_data) and also an empty skeleton(design_table)
    well_data = []
    design_table = []
    subdata = []
    design_subtable = []
    for i in major_axis:
        for j in minor_axis:
            string = i + str(j)
            subdata.append(string)
            design_subtable.append(np.nan)
        well_data.append(subdata)
        design_table.append(design_subtable)
        subdata = []
        design_subtable = []
    well_number = pd.DataFrame(well_data, index=major_axis, columns=minor_axis)
    expt_design_table = pd.DataFrame(design_table, index=major_axis, columns=minor_axis)

    # Specify a writer
    writer = pd.ExcelWriter('design_table.xlsx', engine='xlsxwriter')

    # Write your DataFrame to a file
    r = 0
    for x in range(0, total_reagents):
        expt_design_table.to_excel(writer, 'Sheet1', startrow=r, startcol=0)
        r += rows + 1

    well_number.to_excel(writer, 'Sheet2', startrow=0, startcol=0)

    # Save the result
    writer.save()

    # # The previous part(writing to the excel file was done using xlsxwriter
    # # the next section(formatting the excel file) shall be done using openpyxl
    
    # Load in the workbook
    wb = load_workbook("design_table.xlsx")

    # define the different sheets by name
    sheet2 = wb['Sheet2']
    sheet1 = wb['Sheet1']

    # set the width of the columns
    for i in range(1, columns + 2):
        sheet2.column_dimensions[get_column_letter(i)].width = 4.8
        sheet1.column_dimensions[get_column_letter(i)].width = 4.5
    for i in range(columns + 2, columns + 5):
        sheet1.column_dimensions[get_column_letter(i)].width = 20

    # formatting the table (foreground color and border)
    for i in range(1, columns + 1):
        sheet2[1][i].fill = PatternFill(fgColor="C0C0FF", fill_type="solid")
    for i in range(1, rows + 1):
        sheet2[i+1][0].fill = PatternFill(fgColor="C0C0FF", fill_type="solid")
    set_border(sheet2, get_column_letter(2) + str(2) + ":" + get_column_letter(columns + 1) + str(rows + 1), 'medium')
    r = 0
    for x in range(0, total_reagents):
        sheet1.cell(row=r+1, column=columns + 2).value = "Reagent_" + str(x + 1)
        sheet1.cell(row=r+1, column=columns + 3).value = "source_plate_" + str(x + 1)
        sheet1.cell(row=r+1, column=columns + 4).value = "source_wells_" + str(x + 1)
        for i in range(1, columns + 4):
            sheet1[1+r][i].fill = PatternFill(fgColor="C0C0FF", fill_type="solid")
        for i in range(1, rows + 1):
            sheet1[i+1+r][0].fill = PatternFill(fgColor="C0C0FF", fill_type="solid")
        range_var = get_column_letter(2) + str(r+2) + ":" + get_column_letter(columns + 1) + str(r + rows + 1)
        set_border(sheet1, range_var, 'medium')
        r += rows + 1
    
    # save the file
    wb.save('design_table.xlsx')
    return

# The following function sets borders outside a range of cells in an excel worksheet of desired thickness type
# adopted from an answer from 'stackoverflow' with some minor alterations
def set_border(ws, cell_range, thickness):
    rows = ws[cell_range]
    for row in rows:
        if row == rows[0][0] or row == rows[0][-1] or row == rows[-1][0] or row == rows[-1][-1]:
            pass
        else:
            row[0].border = Border(left=Side(style=thickness))
            row[-1].border = Border(right=Side(style=thickness))
        for c in rows[0]:
            c.border = Border(top=Side(style=thickness))
        for c in rows[-1]:
            c.border = Border(bottom=Side(style=thickness))
    rows[0][0].border = Border(left=Side(style=thickness), top=Side(style=thickness))
    rows[0][-1].border = Border(right=Side(style=thickness), top=Side(style=thickness))
    rows[-1][0].border = Border(left=Side(style=thickness), bottom=Side(style=thickness))
    rows[-1][-1].border = Border(right=Side(style=thickness), bottom=Side(style=thickness))
    return

# calling 'create_file' function
create_file(total_rows, total_columns, reagents)
