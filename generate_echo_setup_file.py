# Install openpyxl, pandas, XlsxWriter, seaborn in your environment or using terminal
# importing the modules
import csv
import os
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import seaborn as sn
# Import `load_workbook` module from `openpyxl`
from openpyxl import load_workbook
# Import `date` module from `datetime`
from datetime import date

##################################################################################
# # The next section in this script needs to be assigned before running the code
##################################################################################

# Assign spreadsheet filename to `design_file`
design_file = '384_well_plate.xlsx'

# dimensions of the plate
# given below for a 384 well plate (16x24)
total_rows = 16
total_columns = 24

# total volume limit of each well in nanolitres
total_volume = 2000
# factor by which volume has been reduced in the Excel design_file
volume_factor = 1000

# Labcyte Echo volume constraints in the source well in nanolitres
max_volume = 65000
min_volume = 20000

# Do you want to calculate volume of water or not?
# Comment the appropriate line among the next two lines based on your requirement
calculate_water = "yes"
# calculate_water = "no"

# Specify plate name and well numbers for water
# this part will not be used if you are not asking the script to calculate water, in that case the values don't matter
water_plate_name: str = "384PP_Plus_AQ_SP"
water_source_well = ['E5', 'E6']

###############################################################
# Actual Code starts: user need not change any part after this
###############################################################

# Defining function to calculate volume of water
def water_volume(well_vol, total_vol):
    return total_vol - well_vol


# Defining function to convert the Excel format to Labcyte Echo format
def echo_format(tot_rows, tot_columns, tot_volume, vol_factor, max_vol, min_vol, file, water_list):
    # to remove "error_file.txt" if it pre-exists in the directory
    if os.path.exists("error_file.txt"):
        os.remove("error_file.txt")
        
    # Load in the workbook
    wb = load_workbook(file)

    # define the different sheets by name
    sheet = wb['Sheet1']
    well = wb['Sheet2']

    # last row in Sheet1
    end = sheet.max_row

    # Calling the function: 'constraints'
    # This returns a list of (list of volume of reagent to be pippeted out of each source well) corresponding to each reagent
    reagent_well, infile_data = constraints(tot_rows, tot_columns, vol_factor, max_vol, min_vol, file)

    # Variables required for switching the source well for the same reagent based on Labcyte Echo volume constraints
    # "reagent volume" will be computed at each iteration so that the values can be compared to the variable "reagent_well" to know when to switch to the next source well
    reagent_volume = []
    flag = []
    i = 1
    # initializing the reagent volume and flag corresponding to all the reagents to 0
    while i < end:
        reagent_volume.append(0)
        flag.append(0)
        i += (tot_rows + 1)
    # adding water to the reagent list if asked by the user to calculate
    if water_list[0] == "yes":
        reagent_volume.append(0)
        flag.append(0)

    # First line of the output design_file
    title = ['Source Plate Name', 'Source Plate Type', 'Source Well', 'Destination Plate Name', 'Destination Well',
             'Transfer Volume', 'Name']
    # format of each new entry:
    # newentry=[1,Plate,Source_well,1,destination_well,volume,Reagent_Name]

    # 2D array for the output design_file
    data = [title]

    # Extracting information in the correct format
    # the k loop is for the rows in the 384 well plate
    # the j loop is for the columns in the 384 well plate
    # the i loop is to scan all the different reagents in the design_file(for the same well)
    # the h loop is for reagent index

    err = 0
    well_volume = 0
    k = 0
    while k < tot_rows:
        j = 2
        while j < tot_columns + 2:
            i = 1
            h = 0
            while i < end:
                if sheet.cell(row=i + k + 1, column=j).value is not None:
                    volume = vol_factor * sheet.cell(row=i + k + 1, column=j).value
                    reagent_volume[h] += sheet.cell(row=i + k + 1, column=j).value

                    n = flag[h] + 1  # index for the well number containing the same reagent
                    if round(reagent_volume[h], 3) > round(sum(reagent_well[h][:n]), 3):
                        flag[h] += 1
                    # this will increase flag only if the computed total volume up to this iteration exceeds the precalculated total volume of the reagent in the source well

                    newentry = [1, sheet.cell(row=i, column=tot_columns + 3).value,
                                sheet.cell(row=i + flag[h], column=tot_columns + 4).value, 1,
                                well.cell(row=2 + k, column=j).value, volume,
                                sheet.cell(row=i, column=tot_columns + 2).value]
                    data.append(newentry)  # adding a new entry to the output variable
                    well_volume += volume
                i += (tot_rows + 1)
                h += 1
            if well_volume > tot_volume:
                warning = 'Total volume in Well Number ' + well.cell(row=2 + k, column=j).value + ' exceeds the maximum limit!'
                error_message(warning)
                print(warning)
                
            # calculating water if specified
            if (water_list[0] == "yes") and (well_volume != 0):
                reagent_volume[-1] += water_volume(well_volume, tot_volume)

                if reagent_volume[-1] > max_vol - min_vol:
                    flag[
                        -1] += 1  # increases the index if total volume of water calculated exceeds the total pipettable volume
                    reagent_volume[-1] = water_volume(well_volume,
                                                      tot_volume)  # initializing the volume of water if shifted to the next source well for water

                # To check if enough wells required for water have been specified
                try:
                    newentry = [1, water_list[1], water_list[2 + flag[-1]], 1, well.cell(row=2 + k, column=j).value,
                                water_volume(well_volume, tot_volume), "water"]
                except IndexError:
                    if err < 1:
                        error_message("Please add more wells in the source plate for water!")
                        print("Please add more wells in the source plate for water!")
                    err += 1
                    water_list.append("")
                    newentry = [1, water_list[1], water_list[2 + flag[-1]], 1, well.cell(row=2 + k, column=j).value,
                                water_volume(well_volume, tot_volume), "water"]
                data.append(newentry)
            j += 1
            well_volume = 0
        k += 1

    if (water_list[0] == "yes"):
        if flag[-1] > 0:
            for i in range(0, flag[-1]):
                infile_data.append(["Water", water_list[2 + i], max_vol / vol_factor])
        infile_data.append(["Water", water_list[2 + flag[-1]], (reagent_volume[-1] + min_vol) / vol_factor])

    # Opening a CSV design_file to enter source plate reagent volumes
    infile = open('reagent_well_volumes.csv', 'w', newline='')
    writer = csv.writer(infile, delimiter=',')
    writer.writerows(infile_data)  # # entering the volume of reagent required in each source well to the CSV design_file
    infile.close()

    # Opening the output design_file in csv format and writing data
    outfile = open('setup_'+date.today().strftime("%d%m%y")+'.csv', 'w', newline='')
    writer = csv.writer(outfile, delimiter=',')
    writer.writerows(data)
    outfile.close()
    return


# Defining Function for checking if adequate volume of reagent has been added to the source wells according to Labcyte Echo volume constraints
def constraints(tot_rows, tot_cols, vol_factor, max_vol, min_vol, file):
    # Load in the workbook
    wb = load_workbook(file)

    # define the different sheets by name
    sheet = wb['Sheet1']

    # last row in Sheet1
    end = sheet.max_row

    #
    in_file_data = [["Reagent", "Source Well Number", "Volume"]]

    # computing volume of each reagent in the source wells
    # the i loop is to scan all the different reagents in the design_file(for the same well)
    # the j loop is for the columns in the 384 well plate
    # the k loop is for the rows in the 384 well plate
    # the h loop is for the row index for each source well in the excel design_file

    reagent_wells = []  # after iterative calculation of volume of reagent required from each source well, the values will be stored in this list variable
    reagent_volume = 0
    i = 1
    while i < end:
        k = 0
        h = 0
        vol = []  # variable for list of volumes for the case of multiple source wells for the same reagent
        while k < tot_rows:
            j = 2
            while j < tot_cols + 2:
                if sheet.cell(row=i + k + 1, column=j).value is not None:
                    volume = vol_factor * sheet.cell(row=i + k + 1, column=j).value
                    reagent_volume += volume
                    if reagent_volume > max_vol - min_vol:
                        v = (reagent_volume - volume + min_vol) / vol_factor
                        infile_dataset = [sheet.cell(row=i + h, column=tot_cols + 2).value,
                                          sheet.cell(row=i + h, column=tot_cols + 4).value, v]
                        in_file_data.append(
                            infile_dataset)  # # storing the volume of reagent required in each source well
                        vol.append((reagent_volume - volume) / vol_factor)
                        reagent_volume = volume
                        # increasing the row index for each source well of the current reagent when the total volume exceeds the total pipettable volume(based on the volume constraints)
                        h += 1
                j += 1
            k += 1
        v = (reagent_volume + min_vol) / vol_factor
        infile_dataset = [sheet.cell(row=i + h, column=tot_cols + 2).value,
                          sheet.cell(row=i + h, column=tot_cols + 4).value, v]
        in_file_data.append(
            infile_dataset)  # # storing the volume of reagent required in the last source well for the current reagent
        vol.append(reagent_volume / vol_factor)

        # warning, if there aren't enough source wells for the current reagent
        if sheet.cell(row=i + h, column=tot_cols + 4).value is None:
            error_message(sheet.cell(row=i, column=tot_cols + 2).value + " needs to be added to more wells in the source plate! Please Check!")
            print(sheet.cell(row=i, column=tot_cols + 2).value + " needs to be added to more wells in the source plate! Please Check!")
        reagent_wells.append(vol)
        reagent_volume = 0
        i += (tot_rows + 1)

    # returning the list "reagent_wells" which is a list of (list of volume of reagent to be pippeted out of each source well) corresponding to each reagent
    return reagent_wells, in_file_data


def input_heatmap(tot_rows, tot_columns, total_vol, vol_factor, file, water_list):
    # Load in the workbook
    wb = load_workbook(file)

    # define the different sheets by name
    sheet = wb['Sheet1']

    # last row in Sheet1
    end = sheet.max_row

    # major and minor axes for the row and column indexes of the plate respectively
    minor_axis = []
    for i in range(1, tot_columns + 1):
        minor_axis.append(i)
    major_axis = []
    alpha = 'A'
    for i in range(1, tot_rows + 1):
        major_axis.append(alpha)
        alpha = chr(ord(alpha) + 1)

    i = 1
    items = []  # list of reagents
    while i < end:
        items.append(sheet.cell(row=i, column=tot_columns + 2).value)
        i += (tot_rows + 1)

    # volume data is the dictionary where each item(reagent) corresponds to a dataframe(volume of reagent to be added to the plate)
    volume_data = {}
    data = []  # 2D array: list of rows
    dataset = []  # 1D array containing values in each row
    h = 0
    i = 1
    while i < end:
        k = 0
        while k < tot_rows:
            j = 2
            while j < tot_columns + 2:
                if sheet.cell(row=i + k + 1, column=j).value is None:
                    dataset.append(np.nan)
                else:
                    dataset.append(sheet.cell(row=i + k + 1, column=j).value)
                j += 1
            data.append(dataset)
            dataset = []
            k += 1
        volume_data[items[h]] = pd.DataFrame(data, index=major_axis, columns=minor_axis)
        data = []
        i += (tot_rows + 1)
        h += 1

    if water_list[0] == "yes":
        water_data = []
        water_dataset = []
        well_volume = 0
        k = 0
        while k < tot_rows:
            j = 1
            while j < tot_columns + 1:
                for m in items:
                    if np.isnan(volume_data[m][j][k]):
                        well_volume += 0
                    else:
                        well_volume += volume_data[m][j][k]
                if well_volume == 0:
                    water_dataset.append(np.nan)
                else:
                    water_dataset.append(water_volume(well_volume, total_vol / vol_factor))
                j += 1
                well_volume = 0
            water_data.append(water_dataset)
            water_dataset = []
            k += 1

        items.append("Water")
        volume_data["Water"] = pd.DataFrame(water_data, index=major_axis, columns=minor_axis)

    # generating a heatmap as a visual representation volumes of each reagent added to the plate
    fig_list = dict()
    for m in items:
        grid_kws = {"height_ratios": (.9, .05), "hspace": .3}
        f, (ax, cbar_ax) = plt.subplots(2, gridspec_kw=grid_kws)
        svm = sn.heatmap(volume_data[m], linewidths=0.5, linecolor='lavender', cmap="YlGnBu", ax=ax, cbar_ax=cbar_ax,
                         cbar_kws={"orientation": "horizontal"})
        figure = svm.get_figure()
        figure.savefig(m + '.png', dpi=400)
        fig_list[m] = m + '.png'

    # returning the volume data dictionary
    return fig_list


# All the error messages to be printed to a TXT design_file
def error_message(error_text):
    error_file = open("error_file.txt", "a")  # append mode
    error_file.write(error_text + "\n")
    error_file.close()
    return

# Calling the Desired Functions
echo_format(total_rows, total_columns, total_volume, volume_factor, max_volume, min_volume,
            design_file, [calculate_water, water_plate_name, *water_source_well])
input_heatmap(total_rows, total_columns, total_volume, volume_factor, design_file,
              [calculate_water, water_plate_name, *water_source_well])
